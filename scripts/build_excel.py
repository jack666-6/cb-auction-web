"""將 twse_auction_raw.json + finmind_cb.json 組成「CB競拍統計.xlsx」。

欄位結構比照原始「CB資料表 的副本.xlsx」裡的 CB競拍 工作表(共64欄)。所有公式都用原始檔案
358筆歷史資料逐欄反推驗證過，誤差在浮點數捨入範圍內：

- A~Z (1~26): 證交所競拍公告原始欄位
- 27 轉換價 / 28 投標結束日現股收盤價 / 29 投標結束日parity%: FinMind + 統一證券CBAS 算出
- 30~33 (合格投標倍數/合格投標筆數/合格投標數量/單筆投標均量): 用 A~Z 既有欄位算出
- 43~46 (最低/加權平均得標成本與溢價率): 直接採用證交所公告的得標價格(不額外加計手續費率)，
  再與 parity% 比較算出溢價率
- 47~63 (掛牌後現股/CB表現): 用「撥券日期」當掛牌第一天，抓母公司現股與CB自己市場價格
  各6個交易日算出；CB還沒掛牌前這些欄位自然是空的
- 得標筆數/得標數量/流標數量/單筆得標均量/法人相關統計(原34~42欄): 這幾欄無法從公開資料
  可靠取得，已依使用者要求整組刪除，不出現在最終欄位裡
- 發行市場/競拍方式/最低每標單投標數量(張)/最高投(得)標數量(張)/每一投標單投標處理費(元)/
  保證金成數(%)/掛牌最高價: 依使用者要求刪除，不出現在最終欄位裡(但twse_auction_raw.json
  裡還是有完整原始資料，只是build_excel.py組表時不寫出這幾欄，所以下次自動抓取不會受影響)
- 所有數值欄位(除了序號、證券代號、日期)都四捨五入到小數點第二位
"""
import json
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent.parent
CACHE_DIR = BASE_DIR / "cache"
RAW_CACHE = CACHE_DIR / "twse_auction_raw.json"
FINMIND_CACHE = CACHE_DIR / "finmind_cb.json"
TCRI_MAP_FILE = BASE_DIR / "data" / "tcri_map.json"
OUTPUT_XLSX = BASE_DIR / "CB競拍統計.xlsx"

HEADERS = [
    "序號", "開標日期", "證券名稱", "證券代號", "發行性質",
    "投標開始日", "投標結束日", "競拍數量(張)", "最低投標價格(元)",
    "撥券日期(上市、上櫃日期)", "主辦券商", "得標總金額(元)", "得標手續費率(%)", "總合格件",
    "合格投標數量(張)", "最低得標價格(元)", "最高得標價格(元)", "得標加權平均價格(元)",
    "承銷價格(元)", "取消競價拍賣(流標或取消)", "轉換價", "投標結束日現股收盤價",
    "投標結束日parity%", "投標結束日前75日股價波動率", "合格投標倍數", "合格投標筆數", "合格投標數量", "單筆投標均量",
    "最低得標成本", "最低得標成本溢價率",
    "得標加權平均成本", "得標加權平均成本溢價率", "掛牌第一天現股收盤價", "掛牌第六天現股收盤價",
    "投標結束日至掛牌第一天現股漲跌幅", "投標結束日至掛牌第六天現股漲跌幅", "掛牌第一天parity%",
    "掛牌第六天parity%", "掛牌第一天開盤價", "掛牌第一天收盤價", "掛牌第六天收盤價",
    "掛牌第一天CB(折)溢價%", "掛牌第六天CB(折)溢價%", "掛牌第一天開盤價最低得標成本報酬率",
    "掛牌第一天開盤價得標加權平均成本報酬率", "掛牌第一天收盤價最低得標成本報酬率",
    "掛牌第一天收盤價得標加權平均成本報酬率", "掛牌第六天最低得標成本報酬率",
    "掛牌第六天得標加權平均成本報酬率",
]

DATE_COLS = {"開標日期", "投標開始日", "投標結束日", "撥券日期(上市、上櫃日期)"}
THOUSANDS_COLS = {
    "競拍數量(張)", "得標總金額(元)", "合格投標數量(張)",
}


def _parse_date(s):
    if not s:
        return None
    return datetime.strptime(s, "%Y-%m-%d").date()


def _r(v, n=6):
    return round(v, n) if v is not None else None


def _fill_derived_columns(r: dict, fm: dict):
    # 30~33: 合格投標統計，全部從既有的 A~Z 欄位算出
    qty_bid = r.get("合格投標數量(張)")
    qty_offer = r.get("競拍數量(張)")
    total_qualified = r.get("總合格件")
    r["合格投標倍數"] = _r(qty_bid / qty_offer) if qty_bid and qty_offer else None
    r["合格投標筆數"] = total_qualified if total_qualified else None
    r["合格投標數量"] = qty_bid if qty_bid else None
    r["單筆投標均量"] = _r(qty_bid / total_qualified) if qty_bid and total_qualified else None

    # 43~46: 得標成本 = 得標價格(直接採用證交所公告數字，不額外加計手續費率)，溢價率 = 成本/parity% - 1
    min_price = r.get("最低得標價格(元)")
    wavg_price = r.get("得標加權平均價格(元)")
    parity = r.get("投標結束日parity%")

    min_cost = wavg_cost = None
    if min_price:
        min_cost = min_price
        r["最低得標成本"] = _r(min_cost)
        r["最低得標成本溢價率"] = _r(min_cost / parity - 1, 8) if parity else None
    if wavg_price:
        wavg_cost = wavg_price
        r["得標加權平均成本"] = _r(wavg_cost)
        r["得標加權平均成本溢價率"] = _r(wavg_cost / parity - 1, 8) if parity else None

    # 47~63: 掛牌後表現，第一天=撥券日期當天，第六天=之後第6個交易日(含第一天)
    stock_seq = fm.get("掛牌現股序列")
    cb_seq = fm.get("掛牌CB序列")
    conv_price = r.get("轉換價")
    bid_end_close = r.get("投標結束日現股收盤價")

    day1_stock = day6_stock = None
    if stock_seq:
        day1_stock = stock_seq[0][2]
        if len(stock_seq) >= 6:
            day6_stock = stock_seq[5][2]
    r["掛牌第一天現股收盤價"] = day1_stock
    r["掛牌第六天現股收盤價"] = day6_stock
    r["投標結束日至掛牌第一天現股漲跌幅"] = (
        _r(day1_stock / bid_end_close - 1, 8) if day1_stock and bid_end_close else None
    )
    r["投標結束日至掛牌第六天現股漲跌幅"] = (
        _r(day6_stock / bid_end_close - 1, 8) if day6_stock and bid_end_close else None
    )

    day1_parity = day1_stock / conv_price * 100 if day1_stock and conv_price else None
    day6_parity = day6_stock / conv_price * 100 if day6_stock and conv_price else None
    r["掛牌第一天parity%"] = _r(day1_parity)
    r["掛牌第六天parity%"] = _r(day6_parity)

    day1_cb_open = day1_cb_close = day6_cb_close = None
    if cb_seq:
        day1_cb_open, day1_cb_close = cb_seq[0][1], cb_seq[0][2]
        if len(cb_seq) >= 6:
            day6_cb_close = cb_seq[5][2]
    r["掛牌第一天開盤價"] = day1_cb_open
    r["掛牌第一天收盤價"] = day1_cb_close
    r["掛牌第六天收盤價"] = day6_cb_close

    r["掛牌第一天CB(折)溢價%"] = (
        _r((day1_cb_close / day1_parity - 1) * 100) if day1_cb_close and day1_parity else None
    )
    r["掛牌第六天CB(折)溢價%"] = (
        _r((day6_cb_close / day6_parity - 1) * 100) if day6_cb_close and day6_parity else None
    )

    r["掛牌第一天開盤價最低得標成本報酬率"] = (
        _r(day1_cb_open / min_cost - 1, 8) if day1_cb_open and min_cost else None
    )
    r["掛牌第一天開盤價得標加權平均成本報酬率"] = (
        _r(day1_cb_open / wavg_cost - 1, 8) if day1_cb_open and wavg_cost else None
    )
    r["掛牌第一天收盤價最低得標成本報酬率"] = (
        _r(day1_cb_close / min_cost - 1, 8) if day1_cb_close and min_cost else None
    )
    r["掛牌第一天收盤價得標加權平均成本報酬率"] = (
        _r(day1_cb_close / wavg_cost - 1, 8) if day1_cb_close and wavg_cost else None
    )
    r["掛牌第六天最低得標成本報酬率"] = (
        _r(day6_cb_close / min_cost - 1, 8) if day6_cb_close and min_cost else None
    )
    r["掛牌第六天得標加權平均成本報酬率"] = (
        _r(day6_cb_close / wavg_cost - 1, 8) if day6_cb_close and wavg_cost else None
    )


def load_rows():
    raw = json.loads(RAW_CACHE.read_text(encoding="utf-8"))
    finmind = json.loads(FINMIND_CACHE.read_text(encoding="utf-8")) if FINMIND_CACHE.exists() else {}
    tcri_map = json.loads(TCRI_MAP_FILE.read_text(encoding="utf-8")) if TCRI_MAP_FILE.exists() else {}

    for r in raw:
        r["開標日期"] = _parse_date(r["開標日期"])
        r["投標開始日"] = _parse_date(r["投標開始日"])
        r["投標結束日"] = _parse_date(r["投標結束日"])
        r["撥券日期(上市、上櫃日期)"] = _parse_date(r["撥券日期(上市、上櫃日期)"])

        cb_code = str(r["證券代號"])
        # 無擔保CB沒有銀行保證，改標示信用評等(TCRI)比「無擔保轉換公司債」這個固定文字更有資訊量；
        # 有擔保的維持原樣，因為擔保銀行才是真正的信用來源，TCRI對這類CB意義不大
        if r.get("發行性質") == "無擔保轉換公司債":
            tcri = tcri_map.get(cb_code[:4])
            if tcri:
                r["發行性質"] = f"TCRI{tcri}"

        fm = finmind.get(cb_code, {})
        conv_price = fm.get("轉換價")
        close_price = fm.get("投標結束日現股收盤價")
        r["轉換價"] = conv_price
        r["投標結束日現股收盤價"] = close_price
        r["投標結束日前75日股價波動率"] = fm.get("投標結束日前75日股價波動率")
        if conv_price and close_price:
            r["投標結束日parity%"] = round(close_price / conv_price * 100, 2)
        else:
            r["投標結束日parity%"] = None

        _fill_derived_columns(r, fm)

    # 依開標日期由舊到新，逐年編序號 YYYY + 3碼序 (與原始工作表慣例一致)
    raw.sort(key=lambda r: (r["開標日期"], str(r["證券代號"])))
    year_seq = {}
    for r in raw:
        y = r["開標日期"].year
        year_seq[y] = year_seq.get(y, 0) + 1
        r["序號"] = f"{y}{year_seq[y]:03d}"

    # 顯示由新到舊排序
    raw.sort(key=lambda r: (r["開標日期"], str(r["證券代號"])), reverse=True)
    return raw


def build_workbook(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CB競拍"

    header_font = Font(size=12)
    header_align = Alignment(wrap_text=True, vertical="center")
    for col, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.alignment = header_align

    for row_idx, r in enumerate(rows, start=2):
        for col_idx, name in enumerate(HEADERS, start=1):
            value = r.get(name)
            if isinstance(value, date):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.number_format = "yyyy/mm/dd"
            else:
                if isinstance(value, float) and name != "證券代號":
                    value = round(value, 2)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if name in THOUSANDS_COLS:
                    cell.number_format = "#,##0"

    ws.freeze_panes = "E2"
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 11.5
    ws.column_dimensions["C"].width = 12.5
    ws.column_dimensions["F"].width = 17
    for col in range(1, len(HEADERS) + 1):
        letter = get_column_letter(col)
        if ws.column_dimensions[letter].width in (None, 0):
            ws.column_dimensions[letter].width = 13

    wb.save(OUTPUT_XLSX)
    print(f"已寫入 {len(rows)} 筆資料到 {OUTPUT_XLSX}")


def main():
    rows = load_rows()
    build_workbook(rows)


if __name__ == "__main__":
    main()
